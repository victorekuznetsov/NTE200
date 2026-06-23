#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Офлайн HTML-дашборд анализа масла для NTE200 и 730E (формат АО Развитие).
Источники: Свод_анализов_масла.xlsm (2025) + Свод_анализов_масла 2026.xlsm (2026).
Самодостаточный HTML: данные встроены JSON, графики на Canvas, без интернета/библиотек."""
import openpyxl, warnings, datetime, json, math, os; warnings.filterwarnings('ignore')
from collections import defaultdict
FILES=['_Данные/Масло/Свод_анализов_масла.xlsm','_Данные/Масло/Свод_анализов_масла 2026.xlsm']
FL={'NTE 200':'NTE200','Komatsu 730E':'730E'}
C={'Fe':'Fe','Cu':'Cu','Cr':'Cr','Pb':'Pb','Sn':'Sn','Al':'Al','Ni':'Ni','Si':'Si',
 'V100':'Вязкость при 100°С (V100),\ncSt (сантистокс)','VI':'Индекс вязкости (VI)',
 'TAN':'Кислотное число (TAN), мгКОН/см3','TBN':'Щелочное число (TBN), мгКОН/см3',
 'OXI':'Окисление\n(OXI), А/0,1мм','NIT':'Нитрование (NIT), А/0,1мм','ST':'Содержание сажи (ST), А/0,1мм',
 'W':'Содержание воды (W), %','F':'Содержание топлива (F), %'}
def num(v):
    try: return round(float(str(v).replace(',','.')),2) if v not in (None,'') else None
    except: return None
def find_header(ws):
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=4,values_only=True)):
        if row and 'Марка' in row and 'Узел' in row:
            return i+1,{h:j for j,h in enumerate(row) if h}
    return None,None
seen=set(); DB=defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
for f in FILES:
    wb=openpyxl.load_workbook(f, read_only=True, data_only=True); ws=wb['Анализы']
    hr,idx=find_header(ws); g=lambda r,n: r[idx[n]] if n in idx and idx[n]<len(r) else None
    for r in ws.iter_rows(min_row=hr+1, values_only=True):
        if not r: continue
        mk=str(g(r,'Марка') or '').strip()
        if mk not in FL: continue
        fl=FL[mk]; mach=str(g(r,'Гаражный №') or '?').strip(); node=str(g(r,'Узел') or '?').strip()
        d=g(r,'Дата отбора пробы'); dt=d.strftime('%Y-%m-%d') if isinstance(d,datetime.datetime) else None
        if not dt: continue
        key=(fl,mach,node,dt,str(g(r,'Номер Протокола испытаний')))
        if key in seen: continue
        seen.add(key)
        p={'dt':dt,'oil':str(g(r,'Марка масла') or '').strip(),'st':str(g(r,'Состояние') or '').strip()}
        for k,col in C.items(): p[k]=num(g(r,col))
        DB[fl][mach][node].append(p)
    wb.close()
for fl in DB:
    for m in DB[fl]:
        for n in DB[fl][m]: DB[fl][m][n].sort(key=lambda x:x['dt'])
# percentiles per fleet+node+param
def pctl(a,q):
    a=sorted(a);
    if not a: return None
    k=(len(a)-1)*q; f=math.floor(k); c=math.ceil(k); return round(a[f] if f==c else a[f]+(a[c]-a[f])*(k-f),2)
PARAMS=['Fe','Cu','Cr','Pb','Al','Si','Sn','Ni','TBN','TAN','V100','VI','OXI','NIT','ST','W','F']
PCT=defaultdict(dict)
for fl in DB:
    nodevals=defaultdict(lambda: defaultdict(list))
    for m in DB[fl]:
        for n in DB[fl][m]:
            for p in DB[fl][m][n]:
                for k in PARAMS:
                    if p.get(k) is not None: nodevals[n][k].append(p[k])
    for n in nodevals:
        for k in PARAMS:
            v=nodevals[n][k]
            if v: PCT[f"{fl}||{n}"][k]={'p50':pctl(v,.5),'p90':pctl(v,.9),'p95':pctl(v,.95)}
META={'gen':datetime.date.today().isoformat(),
      'fleets':{fl:{'machines':len(DB[fl]),'probes':sum(len(DB[fl][m][n]) for m in DB[fl] for n in DB[fl][m])} for fl in DB}}
# Воздействия (ремонты ГБЦ/клапанов) из техотчётов — метки для графика (узел ДВС)
EVENTS={'NTE200':{
 '51':[['2025-07-17','Замена ГБЦ']],
 '58':[['2025-08-10','Толкатель/клапан'],['2026-03-16','Клапан ГБЦ']],
 '47':[['2025-08-11','Замена ГБЦ']],
 '48':[['2025-12-17','Замена ГБЦ'],['2026-03-31','Повторный ремонт']],
 '43':[['2026-01-21','Клапаны+форсунки']],
 '83':[['2026-03-30','Замена ГБЦ']],
 '55':[['2026-05-13','Клапаны']],
 '78':[['2026-05-20','Клапаны']],
 '69':[['2026-05-25','Клапаны']],
 '54':[['2026-06-12','Клапаны']],
 '89':[['2026-06-12','Клапаны']],
 '77':[['2026-06-16','Втулка']],
 '87':[['2026-06-18','Клапаны']],
}}
data_json=json.dumps({'DB':DB,'PCT':PCT,'META':META,'PARAMS':PARAMS,'EVENTS':EVENTS},ensure_ascii=False,separators=(',',':'))

HTML=r'''<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8">
<title>Дашборд масла — NTE200 / 730E · АО Развитие</title>
<style>
:root{--bg:#F5F5F5;--ink:#293136;--acc:#3EF0AF;--red:#E05252;--org:#E08C2E;--blu:#4EC9F0;--gry:#707880;--card:#FFFFFF;--line:#D7DCDF}
*{box-sizing:border-box;font-family:Calibri,Segoe UI,Arial,sans-serif}
body{margin:0;background:var(--bg);color:var(--ink)}
header{background:var(--ink);color:#fff;padding:10px 18px;display:flex;align-items:center;gap:18px;border-bottom:4px solid var(--acc)}
header h1{font-size:18px;margin:0}
.tg{display:flex;gap:6px;margin-left:auto}
.tg button{background:#3a454c;color:#cfd6da;border:0;padding:7px 16px;border-radius:6px;cursor:pointer;font-weight:bold}
.tg button.act{background:var(--acc);color:var(--ink)}
.kpis{display:flex;gap:12px;padding:12px 18px;flex-wrap:wrap}
.kpi{background:var(--card);border-radius:8px;padding:10px 16px;min-width:130px;border-top:3px solid var(--acc)}
.kpi .v{font-size:22px;font-weight:bold}.kpi .l{font-size:12px;color:var(--gry)}
.wrap{display:flex;gap:12px;padding:0 18px 18px}
.side{width:120px;background:var(--card);border-radius:8px;padding:8px;max-height:74vh;overflow:auto}
.side b{font-size:12px;color:var(--gry)}
.mch{padding:5px 8px;margin:2px 0;border-radius:5px;cursor:pointer;font-size:13px}
.mch:hover{background:#eef1f2}.mch.act{background:var(--acc);font-weight:bold}
.main{flex:1;display:flex;flex-direction:column;gap:12px}
.bar{display:flex;gap:6px;flex-wrap:wrap;align-items:center}
.bar .lbl{font-size:12px;color:var(--gry);font-weight:bold;margin-right:4px}
.pc{background:var(--card);border:1px solid var(--line);padding:4px 11px;border-radius:14px;cursor:pointer;font-size:13px}
.pc.act{background:var(--ink);color:#fff;border-color:var(--ink)}
.panel{background:var(--card);border-radius:8px;padding:12px}
canvas{width:100%;height:340px}
table{width:100%;border-collapse:collapse;font-size:12px}
th,td{padding:4px 7px;border-bottom:1px solid var(--line);text-align:right}
th:first-child,td:first-child{text-align:left}
th{color:var(--gry);font-weight:bold}
.crit{color:var(--red);font-weight:bold}.warn{color:var(--org);font-weight:bold}
.foot{color:var(--gry);font-size:11px;padding:6px 18px}
.tabs{display:flex;gap:6px;margin-bottom:8px}
.tabs button{background:transparent;border:1px solid var(--line);padding:5px 12px;border-radius:6px;cursor:pointer;font-size:13px}
.tabs button.act{background:var(--acc);border-color:var(--acc);font-weight:bold}
</style></head><body>
<header><h1>🛢️ Дашборд анализа масла — NTE200 / 730E</h1>
<span style="color:#cfd6da;font-size:12px">АО Развитие · Полюс Магадан</span>
<div class="tg" id="tg"></div></header>
<div class="kpis" id="kpis"></div>
<div class="bar" style="padding:0 18px 8px"><span class="lbl">УЗЕЛ:</span><span id="nodes"></span>
<span class="lbl" style="margin-left:16px">ПАРАМЕТР:</span><span id="params"></span></div>
<div class="wrap">
<div class="side" id="side"></div>
<div class="main">
 <div class="panel"><div id="chTitle" style="font-weight:bold;margin-bottom:6px"></div><canvas id="cv"></canvas>
 <div style="font-size:11px;color:var(--gry);margin-top:4px">Линии: P50 (норма, зелёная) · P90 (внимание, оранжевая) · P95 (критично, красная) — по парку+узлу.</div></div>
 <div class="panel">
   <div class="tabs"><button class="act" onclick="setTab('tbl',this)">Пробы машины</button>
   <button onclick="setTab('crit',this)">Критичные по парку</button>
   <button onclick="setTab('rank',this)">Рейтинг износа (ДВС)</button></div>
   <div id="tab"></div>
 </div>
</div></div>
<div class="foot">Сгенерировано из Свод_анализов_масла.xlsm (2025) + Свод_анализов_масла 2026.xlsm (2026). Дата: __GEN__ · Конфиденциально</div>
<script>
const D=__DATA__;
const PLBL={Fe:'Fe ppm',Cu:'Cu ppm',Cr:'Cr ppm',Pb:'Pb ppm',Al:'Al ppm',Si:'Si ppm',Sn:'Sn ppm',Ni:'Ni ppm',
TBN:'TBN мгКОН',TAN:'TAN мгКОН',V100:'V100 cSt',VI:'VI',OXI:'OXI',NIT:'NIT',ST:'Сажа ST',W:'Вода %',F:'Топливо %'};
let cF='NTE200',cM=null,cN='ДВС',cP='Fe',cTab='tbl';
const $=id=>document.getElementById(id);
function machines(){return Object.keys(D.DB[cF]).sort((a,b)=>(parseInt(a)||0)-(parseInt(b)||0))}
function nodesOf(){let s=new Set();for(const m in D.DB[cF])for(const n in D.DB[cF][m])s.add(n);return [...s]}
function init(){
 $('tg').innerHTML=['NTE200','730E'].map(f=>`<button class="${f==cF?'act':''}" onclick="selF('${f}')">${f}</button>`).join('');
 const me=D.META.fleets[cF];
 let crit=critList().length;
 $('kpis').innerHTML=`<div class="kpi"><div class="v">${me.machines}</div><div class="l">машин</div></div>
 <div class="kpi"><div class="v">${me.probes}</div><div class="l">проб масла</div></div>
 <div class="kpi"><div class="v" style="color:var(--red)">${crit}</div><div class="l">критичных проб (&gt;P95)</div></div>
 <div class="kpi"><div class="v">${nodesOf().length}</div><div class="l">узлов</div></div>`;
 $('nodes').innerHTML=nodesOf().map(n=>`<span class="pc ${n==cN?'act':''}" onclick="selN('${n}')">${n}</span>`).join(' ');
 $('params').innerHTML=D.PARAMS.map(p=>`<span class="pc ${p==cP?'act':''}" onclick="selP('${p}')">${PLBL[p]||p}</span>`).join(' ');
 $('side').innerHTML='<b>МАШИНЫ</b>'+machines().map(m=>`<div class="mch ${m==cM?'act':''}" onclick="selM('${m}')">№${m}</div>`).join('');
 draw();renderTab();
}
function selF(f){cF=f;cM=null;if(!nodesOf().includes(cN))cN='ДВС';init()}
function selN(n){cN=n;init()}
function selP(p){cP=p;init()}
function selM(m){cM=m;cTab='tbl';init()}
function setTab(t,b){cTab=t;document.querySelectorAll('.tabs button').forEach(x=>x.classList.remove('act'));b.classList.add('act');renderTab()}
function series(){if(!cM||!D.DB[cF][cM][cN])return [];return D.DB[cF][cM][cN].filter(p=>p[cP]!=null).map(p=>({x:p.dt,y:p[cP]}))}
function pct(){return (D.PCT[cF+'||'+cN]||{})[cP]||{}}
function critList(){let out=[];const node=cN;for(const m in D.DB[cF]){for(const p of (D.DB[cF][m][node]||[])){const pc=(D.PCT[cF+'||'+node]||{});let fl=[];for(const k of ['Fe','Cu','Si','ST','OXI']){if(pc[k]&&p[k]!=null&&p[k]>pc[k].p95)fl.push(k)}if(fl.length)out.push({m,dt:p.dt,fl,Fe:p.Fe,Si:p.Si,ST:p.ST})}}return out.sort((a,b)=>b.dt<a.dt?-1:1)}
function evOf(){if(cF!=='NTE200'||cN!=='ДВС'||!cM)return [];return ((D.EVENTS.NTE200||{})[cM]||[])}
function ts(d){return new Date(d).getTime()}
function fmt(t){const z=new Date(t);return ('0'+z.getDate()).slice(-2)+'.'+('0'+(z.getMonth()+1)).slice(-2)+'.'+String(z.getFullYear()).slice(2)}
function draw(){
 const cv=$('cv');const dpr=devicePixelRatio||1;const W=cv.clientWidth,H=340;cv.width=W*dpr;cv.height=H*dpr;
 const x=cv.getContext('2d');x.scale(dpr,dpr);x.clearRect(0,0,W,H);
 $('chTitle').textContent=cM?`№${cM} · ${cN} · ${PLBL[cP]||cP}`:'Выберите машину слева';
 const s=series().map(d=>({t:ts(d.x),y:d.y,dt:d.x}));const pc=pct();const ev=evOf();
 const pad={l:46,r:14,t:14,b:46};const pw=W-pad.l-pad.r,ph=H-pad.t-pad.b;
 let ys=s.map(d=>d.y);['p50','p90','p95'].forEach(k=>{if(pc[k]!=null)ys.push(pc[k])});
 if(!ys.length)ys=[0,1];let ymax=Math.max(...ys)*1.1||1,ymin=Math.min(0,...ys);
 // time scale: include event dates in range
 let tvals=s.map(d=>d.t).concat(ev.map(e=>ts(e[0])));
 if(!tvals.length){x.fillStyle='#707880';x.font='13px Calibri';x.fillText('Нет данных по выбору',pad.l+10,pad.t+30);return}
 let tmin=Math.min(...tvals),tmax=Math.max(...tvals);if(tmin===tmax){tmin-=864e5;tmax+=864e5}
 const X=t=>pad.l+pw*(t-tmin)/(tmax-tmin);const Y=v=>pad.t+ph-ph*(v-ymin)/(ymax-ymin||1);
 // axes
 x.strokeStyle='#D7DCDF';x.lineWidth=1;x.beginPath();x.moveTo(pad.l,pad.t);x.lineTo(pad.l,pad.t+ph);x.lineTo(pad.l+pw,pad.t+ph);x.stroke();
 x.fillStyle='#707880';x.font='11px Calibri';
 for(let g=0;g<=4;g++){const v=ymin+(ymax-ymin)*g/4;const yy=Y(v);x.fillText(v.toFixed(1),4,yy+3);x.strokeStyle='#EEF1F2';x.beginPath();x.moveTo(pad.l,yy);x.lineTo(pad.l+pw,yy);x.stroke()}
 // date ticks on X
 x.fillStyle='#707880';x.textAlign='center';
 for(let g=0;g<=5;g++){const t=tmin+(tmax-tmin)*g/5;const xx=X(t);x.strokeStyle='#EEF1F2';x.beginPath();x.moveTo(xx,pad.t);x.lineTo(xx,pad.t+ph);x.stroke();x.fillStyle='#707880';x.fillText(fmt(t),xx,pad.t+ph+14)}
 x.textAlign='left';
 // percentile lines
 [['p50','#2EA043'],['p90','#E08C2E'],['p95','#E05252']].forEach(([k,c])=>{if(pc[k]==null)return;x.strokeStyle=c;x.setLineDash([5,4]);x.beginPath();x.moveTo(pad.l,Y(pc[k]));x.lineTo(pad.l+pw,Y(pc[k]));x.stroke();x.setLineDash([]);x.fillStyle=c;x.fillText(k.toUpperCase()+' '+pc[k],pad.l+pw-52,Y(pc[k])-3)});
 // event markers (воздействия)
 ev.forEach(e=>{const xx=X(ts(e[0]));x.strokeStyle='#9B59B6';x.setLineDash([3,3]);x.lineWidth=1.5;x.beginPath();x.moveTo(xx,pad.t);x.lineTo(xx,pad.t+ph);x.stroke();x.setLineDash([]);
  x.save();x.translate(xx,pad.t+2);x.rotate(-Math.PI/2);x.fillStyle='#9B59B6';x.font='10px Calibri';x.textAlign='right';x.fillText('▲ '+e[1]+' '+fmt(ts(e[0])),-2,10);x.restore();x.textAlign='left'});
 // series
 if(s.length){x.strokeStyle='#293136';x.lineWidth=2;x.beginPath();s.forEach((d,i)=>{i?x.lineTo(X(d.t),Y(d.y)):x.moveTo(X(d.t),Y(d.y))});x.stroke();
  s.forEach(d=>{const crit=pc.p95!=null&&d.y>pc.p95;x.fillStyle=crit?'#E05252':'#293136';x.beginPath();x.arc(X(d.t),Y(d.y),crit?4:3,0,7);x.fill()});}
}
function renderTab(){
 let h='';
 if(cTab=='tbl'){
  if(!cM){h='<div style="color:#707880">Выберите машину.</div>'}else{
   const all=(D.DB[cF][cM][cN]||[]).slice().reverse();const rows=all;const pc=pct();
   h=`<div style="font-size:12px;color:#707880;margin-bottom:4px">Все пробы (${all.length}) с начала отбора. <span style="color:#9B59B6">▲ фиолетовые метки на графике — воздействия (ремонты ГБЦ/клапанов) из техотчётов.</span></div>`;
   h+='<div style="max-height:300px;overflow:auto"><table><tr><th>Дата</th><th>Масло</th><th>Fe</th><th>Cu</th><th>Si</th><th>TBN</th><th>Сажа</th><th>V100</th><th>Состояние</th></tr>';
   for(const p of rows){const c=(k)=>{const pcc=(D.PCT[cF+'||'+cN]||{})[k];if(pcc&&p[k]!=null&&p[k]>pcc.p95)return'crit';if(pcc&&p[k]!=null&&p[k]>pcc.p90)return'warn';return''};
    h+=`<tr><td>${p.dt}</td><td style="text-align:left">${p.oil||''}</td><td class="${c('Fe')}">${p.Fe??''}</td><td class="${c('Cu')}">${p.Cu??''}</td><td class="${c('Si')}">${p.Si??''}</td><td>${p.TBN??''}</td><td class="${c('ST')}">${p.ST??''}</td><td>${p.V100??''}</td><td style="text-align:left">${p.st||''}</td></tr>`}
   h+='</table></div>'}
 }else if(cTab=='crit'){
  const cl=critList().slice(0,40);
  h=`<div style="font-size:12px;color:#707880;margin-bottom:6px">Узел ${cN}: пробы с превышением P95 по Fe/Cu/Si/Сажа/OXI</div><table><tr><th>Маш</th><th>Дата</th><th>Превышения</th><th>Fe</th><th>Si</th><th>Сажа</th></tr>`;
  for(const r of cl)h+=`<tr><td>№${r.m}</td><td>${r.dt}</td><td style="text-align:left" class="crit">${r.fl.join(', ')}</td><td>${r.Fe??''}</td><td>${r.Si??''}</td><td>${r.ST??''}</td></tr>`;
  h+='</table>'+(cl.length?'':'<div style="color:#2EA043">Нет критичных проб по узлу.</div>');
 }else{
  const med=(arr)=>{arr=arr.filter(v=>v!=null).sort((a,b)=>a-b);return arr.length?arr[Math.floor(arr.length/2)]:null};
  let rows=machines().map(m=>{const pr=D.DB[cF][m]['ДВС']||[];return {m,n:pr.length,Fe:med(pr.map(p=>p.Fe)),TBN:med(pr.map(p=>p.TBN)),ST:med(pr.map(p=>p.ST))}}).filter(r=>r.n).sort((a,b)=>(b.Fe||0)-(a.Fe||0));
  h='<div style="font-size:12px;color:#707880;margin-bottom:6px">Машины по медианному Fe на ДВС (выше = больше износ)</div><table><tr><th>Маш</th><th>Проб</th><th>Fe med</th><th>TBN med</th><th>Сажа med</th></tr>';
  for(const r of rows)h+=`<tr><td>№${r.m}</td><td>${r.n}</td><td>${r.Fe??''}</td><td>${r.TBN??''}</td><td>${r.ST??''}</td></tr>`;h+='</table>';
 }
 $('tab').innerHTML=h;
}
window.addEventListener('resize',draw);init();
</script></body></html>'''
HTML=HTML.replace('__DATA__',data_json).replace('__GEN__',META['gen'])
out="05 Отчёты/Аналитические/Дашборд масла NTE200 730E.html"
open(out,'w',encoding='utf-8').write(HTML)
print("saved:",out,"|",os.path.getsize(out)//1024,"KB | проб:",sum(v['probes'] for v in META['fleets'].values()))

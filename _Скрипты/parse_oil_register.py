#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Расширенный анализ масла ДВС для NTE200 и 730E.
Источники: Свод_анализов_масла.xlsm (2025 и ранее) + Свод_анализов_масла 2026.xlsm (2026, обновляемый).
Устойчивое определение строки заголовка. Дедуп по (парк,машина,узел,дата,протокол)."""
import openpyxl, warnings, datetime, statistics, math, json, sys; warnings.filterwarnings('ignore')
from collections import defaultdict, Counter
FILES=['_Данные/Масло/Свод_анализов_масла.xlsm','_Данные/Масло/Свод_анализов_масла 2026.xlsm']
FL={'NTE 200':'NTE200','Komatsu 730E':'730E'}
KEYS={'V100':'Вязкость при 100°С (V100),\ncSt (сантистокс)','VI':'Индекс вязкости (VI)',
 'TAN':'Кислотное число (TAN), мгКОН/см3','TBN':'Щелочное число (TBN), мгКОН/см3',
 'OXI':'Окисление\n(OXI), А/0,1мм','Вода%':'Содержание воды (W), %','Топливо%':'Содержание топлива (F), %',
 'Сажа':'Содержание сажи (ST), А/0,1мм'}
def num(v):
    try: return float(str(v).replace(',','.')) if v not in (None,'') else None
    except: return None
def pct(a,q):
    a=sorted(a); k=(len(a)-1)*q; f=math.floor(k); c=math.ceil(k); return a[f] if f==c else a[f]+(a[c]-a[f])*(k-f)
def find_header(ws):
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=4,values_only=True)):
        if row and any(c=='Марка' for c in row) and any(c=='Узел' for c in row):
            return i+1,{h:j for j,h in enumerate(row) if h}
    return None,None
seen=set(); eng=[]; bynode=defaultdict(lambda: defaultdict(lambda: defaultdict(list))); allds=defaultdict(list)
for f in FILES:
    wb=openpyxl.load_workbook(f, read_only=True, data_only=True); ws=wb['Анализы']
    hr,idx=find_header(ws)
    g=lambda r,n: r[idx[n]] if n in idx and idx[n]<len(r) else None
    for r in ws.iter_rows(min_row=hr+1, values_only=True):
        if not r: continue
        mk=str(g(r,'Марка') or '').strip()
        if mk not in FL: continue
        fl=FL[mk]; node=str(g(r,'Узел') or '').strip()
        key=(fl,str(g(r,'Гаражный №')),node,str(g(r,'Дата отбора пробы')),str(g(r,'Номер Протокола испытаний')))
        if key in seen: continue
        seen.add(key)
        d=g(r,'Дата отбора пробы')
        if isinstance(d,datetime.datetime): allds[fl].append(d)
        for p in ['Fe','Cu','Cr','Pb','Al','Si']:
            v=num(g(r,p))
            if v is not None: bynode[fl][node][p].append(v)
        if node=='ДВС':
            rec={'Дата':d.isoformat() if isinstance(d,datetime.datetime) else None,
                 'Производитель':str(g(r,'Производитель масла') or '?').strip(),
                 'Марка масла':str(g(r,'Марка масла') or '?').strip(),
                 'Вязкость':str(g(r,'Вязкость') or '').strip(),
                 'Тип':str(g(r,'Тип масла') or '?').strip(),
                 'Fe':num(g(r,'Fe'))}
            for k,col in KEYS.items(): rec[k]=num(g(r,col))
            eng.append((fl,rec))
    wb.close()
json.dump(eng, open('/tmp/oil_engine.json','w'), ensure_ascii=False)
print("ДВС проб:", Counter(fl for fl,_ in eng), "| период:",
      min(min(allds['NTE200']),min(allds['730E'])).date(),"—",max(max(allds['NTE200']),max(allds['730E'])).date())
json.dump({fl:{n:{p:bynode[fl][n][p] for p in bynode[fl][n]} for n in bynode[fl]} for fl in bynode},
          open('/tmp/oil_nodes.json','w'), ensure_ascii=False)
json.dump({fl:[d.isoformat() for d in allds[fl]] for fl in allds}, open('/tmp/oil_ds.json','w'), ensure_ascii=False)

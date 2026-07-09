# Пересобирает данные для "Дашборд весового контроля NTE200.html" из ВСЕХ доступных
# выгрузок весового устройства: файлов Весовая_*.xlsx в корне репозитория и более
# старого снэпшота, упакованного в _Архивы/Payload NTE200.7z (папка 05.05.2026).
#
# Экспорт с весового контроллера — скользящее окно (не полная история с начала
# эксплуатации), поэтому у одного борта может быть несколько выгрузок за разные даты
# с пересекающимися диапазонами PayloadNumber/времени. Чтобы не задвоить циклы при
# объединении всех источников, каждая строка приводится к natural key
# (TruckId, Год, Месяц, День, Час, Минута, Секунда, PayloadNumber) — так одинаковые
# физические циклы, встречающиеся в двух выгрузках, схлопываются в одну запись
# (проверено на борту №71: 1973 общих цикла между майским архивом и июльским
# корневым файлом дали 0 расхождений в FinalPayload/LoadPercentage).
#
# Использование: python3 build_weighing_dashboard_data.py, затем вставить содержимое
# weighing_data.json в HTML-дашборд как значение константы DATA.
import openpyxl, glob, re, json, os, io, tempfile, statistics
from collections import defaultdict

try:
    import py7zr
except ImportError:
    py7zr = None

RATED_T = 178.0  # nominal payload, derived from FinalPayload/LoadPercentage across all trucks (median 1780 -> /10 = 178.0 t)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MONTH_NAMES = {8: 'авг', 9: 'сен', 10: 'окт', 11: 'ноя', 12: 'дек', 1: 'янв', 2: 'фев', 3: 'мар', 4: 'апр', 5: 'май', 6: 'июн', 7: 'июл'}
BIN_WIDTH = 10  # tonnes


def bin_key(t):
    return int(t // BIN_WIDTH) * BIN_WIDTH


def iter_source_workbooks():
    """Yields (source_label, truck_id_from_filename, openpyxl_workbook) for every
    Весовая_*.xlsx found at repo root and inside _Архивы/Payload NTE200.7z."""
    for path in sorted(glob.glob(os.path.join(REPO_ROOT, 'Весовая_*.xlsx'))):
        fname = os.path.basename(path)
        m = re.match(r'Весовая_(\d+)_', fname)
        if not m:
            continue
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        yield fname, m.group(1), wb
        wb.close()

    archive_path = os.path.join(REPO_ROOT, '_Архивы', 'Payload NTE200.7z')
    if py7zr and os.path.exists(archive_path):
        with tempfile.TemporaryDirectory() as tmp:
            with py7zr.SevenZipFile(archive_path, mode='r') as z:
                names = [n for n in z.getnames() if n.lower().endswith('.xlsx')]
                z.extract(path=tmp, targets=names)
            for name in names:
                fname = os.path.basename(name)
                m = re.match(r'Весовая_(\d+)_', fname)
                if not m:
                    continue
                wb = openpyxl.load_workbook(os.path.join(tmp, name), data_only=True, read_only=True)
                yield f'archive:{fname}', m.group(1), wb
                wb.close()


# ---- pass 1: collect every row from every source, deduped by physical cycle key ----
unique_rows = {}  # key -> (truck_id, row_tuple, idx)
row_idx = None
sources_seen = defaultdict(set)  # truck_id -> set of source labels contributing rows

for source_label, truck_from_name, wb in iter_source_workbooks():
    sn = [s for s in wb.sheetnames if 'Payload' in s] or wb.sheetnames
    ws = None
    for candidate in sn:
        probe = wb[candidate]
        first_row = next(probe.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if first_row and 'TruckId' in first_row:
            ws = probe
            break
    if ws is None:
        continue
    rows = ws.iter_rows(min_row=2, values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    row_idx = idx  # schema identical across all source files

    n_new = 0
    for r in rows:
        fp = r[idx['FinalPayload']]
        lp = r[idx['LoadPercentage']]
        if not fp or not lp:
            continue
        truck_id = str(r[idx['TruckId']])
        key = (truck_id, r[idx['ear']], r[idx['Month']], r[idx['Day']],
               r[idx['Hour']], r[idx['Minute']], r[idx['Second']], r[idx['PayloadNumber']])
        if key not in unique_rows:
            unique_rows[key] = r
            n_new += 1
    sources_seen[truck_from_name].add(f'{source_label} (+{n_new} новых циклов)')

idx = row_idx
print(f'Уникальных циклов после дедупликации: {len(unique_rows)}')

# ---- pass 2: group deduped rows by truck ----
by_truck = defaultdict(list)
for (truck_id, *_), r in unique_rows.items():
    by_truck[truck_id].append(r)

trucks = {}
fleet_load_hist = defaultdict(int)
fleet_month_cycles = defaultdict(lambda: defaultdict(int))
fleet_month_loadsum = defaultdict(lambda: defaultdict(float))

for truck in sorted(by_truck.keys(), key=int):
    rows = by_truck[truck]
    n = 0
    payloads, loadpcts = [], []
    stop_loaded, moving_loaded, moving_empty, stop_empty, loading_t, dumping_t = [], [], [], [], [], []
    dist_empty, dist_loaded, ld_speed, em_speed, lr_tkph, rr_tkph, fuel = [], [], [], [], [], [], []
    hist = defaultdict(int)
    first_dt = last_dt = None

    for r in rows:
        fp = r[idx['FinalPayload']]
        lp = r[idx['LoadPercentage']]
        n += 1
        t = fp / 10.0
        payloads.append(t)
        loadpcts.append(lp)
        hist[bin_key(t)] += 1
        fleet_load_hist[bin_key(t)] += 1

        yr, mo, d, h, mi, s = r[idx['ear']], r[idx['Month']], r[idx['Day']], r[idx['Hour']], r[idx['Minute']], r[idx['Second']]
        full_year = 2000 + yr
        dt = (full_year, mo, d, h, mi, s)
        if first_dt is None or dt < first_dt:
            first_dt = dt
        if last_dt is None or dt > last_dt:
            last_dt = dt
        fleet_month_cycles[(full_year, mo)][truck] += 1
        fleet_month_loadsum[(full_year, mo)][truck] += lp

        sl, ml, me, se = r[idx['StopLoadedTime']], r[idx['MovingLoadedTime']], r[idx['MovingEmptyTime']], r[idx['StopEmptyTime']]
        lt, dt_ = r[idx['LoadingTime']], r[idx['DumpingTime']]
        if sl is not None: stop_loaded.append(sl)
        if ml is not None: moving_loaded.append(ml)
        if me is not None: moving_empty.append(me)
        if se is not None: stop_empty.append(se)
        if lt is not None: loading_t.append(lt)
        if dt_ is not None: dumping_t.append(dt_)

        de, dl = r[idx['DistanceEmpty']], r[idx['DistanceLoaded']]
        if de: dist_empty.append(de)
        if dl: dist_loaded.append(dl)

        lds, ems = r[idx['LdMaxSpeed']], r[idx['EmMaxSpeed']]
        if lds: ld_speed.append(lds / 1000.0)
        if ems: em_speed.append(ems / 1000.0)

        lr, rr = r[idx['LR_TKPH']], r[idx['RR_TKPH']]
        if lr is not None: lr_tkph.append(lr / 10.0)
        if rr is not None: rr_tkph.append(rr / 10.0)

        fl = r[idx['FuelLevel']]
        if fl is not None: fuel.append(fl)

    if n == 0:
        continue

    avg = lambda L: (sum(L) / len(L)) if L else 0
    overload = sum(1 for lp in loadpcts if lp > 110)
    target = sum(1 for lp in loadpcts if 90 <= lp <= 110)
    underload = sum(1 for lp in loadpcts if lp < 90)

    cyc_stop_loaded = avg(stop_loaded) / 60.0
    cyc_moving_loaded = avg(moving_loaded) / 60.0
    cyc_moving_empty = avg(moving_empty) / 60.0
    cyc_stop_empty = avg(stop_empty) / 60.0
    cyc_loading = avg(loading_t) / 60.0
    cyc_dumping = avg(dumping_t) / 60.0
    total_cycle_min = cyc_stop_loaded + cyc_moving_loaded + cyc_moving_empty + cyc_stop_empty + cyc_loading + cyc_dumping

    total_tonnage = sum(payloads)
    productivity_tph = (avg(payloads) / (total_cycle_min / 60.0)) if total_cycle_min else 0

    trucks[truck] = {
        'truck': truck,
        'cycles': n,
        'avg_payload_t': round(avg(payloads), 1),
        'avg_load_pct': round(avg(loadpcts), 1),
        'total_tonnage_t': round(total_tonnage, 0),
        'overload_pct': round(100 * overload / n, 1),
        'target_pct': round(100 * target / n, 1),
        'underload_pct': round(100 * underload / n, 1),
        'hist': dict(sorted(hist.items())),
        'cycle_min': {
            'stop_loaded': round(cyc_stop_loaded, 2),
            'moving_loaded': round(cyc_moving_loaded, 2),
            'moving_empty': round(cyc_moving_empty, 2),
            'stop_empty': round(cyc_stop_empty, 2),
            'loading': round(cyc_loading, 2),
            'dumping': round(cyc_dumping, 2),
            'total': round(total_cycle_min, 2),
        },
        'dist_loaded_km': round(avg(dist_loaded) / 1000.0, 2),
        'dist_empty_km': round(avg(dist_empty) / 1000.0, 2),
        'ld_speed_kmh': round(avg(ld_speed), 1),
        'em_speed_kmh': round(avg(em_speed), 1),
        'lr_tkph': round(avg(lr_tkph), 1),
        'rr_tkph': round(avg(rr_tkph), 1),
        'fuel_pct': round(avg(fuel), 1),
        'productivity_tph': round(productivity_tph, 1),
    }
    y_first, mo_first, d_first = first_dt[0], first_dt[1], first_dt[2]
    y_last, mo_last, d_last = last_dt[0], last_dt[1], last_dt[2]
    trucks[truck]['period_str'] = f"{d_first:02d}.{mo_first:02d}.{y_first} – {d_last:02d}.{mo_last:02d}.{y_last}"

fleet = {
    'rated_t': RATED_T,
    'trucks_n': len(trucks),
    'cycles_total': sum(t['cycles'] for t in trucks.values()),
    'tonnage_total': round(sum(t['total_tonnage_t'] for t in trucks.values()), 0),
    'load_hist': dict(sorted(fleet_load_hist.items())),
    'month_trend': [
        {
            'label': f"{MONTH_NAMES.get(mo, mo)} {y}",
            'cycles': sum(fleet_month_cycles[(y, mo)].values()),
            'avg_load_pct': round(sum(fleet_month_loadsum[(y, mo)].values()) / sum(fleet_month_cycles[(y, mo)].values()), 1) if sum(fleet_month_cycles[(y, mo)].values()) else 0
        } for (y, mo) in sorted(fleet_month_cycles.keys())
    ]
}

out = {'fleet': fleet, 'trucks': trucks}
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'weighing_data.json')
with open(out_path, 'w', encoding='utf-8') as fh:
    json.dump(out, fh, ensure_ascii=False, indent=1)

print('trucks:', len(trucks))
print('fleet cycles', fleet['cycles_total'], 'tonnage', fleet['tonnage_total'])
print('month trend', fleet['month_trend'])
